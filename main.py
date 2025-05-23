import os
import logging
import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# pip install pandas requests beautifulsoup4 openpyxl
# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

# Setup a session with a realistic User-Agent
session = requests.Session()
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
})

# Red fill for missing items
def get_red_fill():
    return PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')


def find_product_url(sku: str) -> str:
    """
    Search for the given SKU on praktis.bg via the catalogsearch endpoint.
    Returns full product URL if found, or "Item doesn't exist" otherwise.
    """
    search_url = f"https://praktis.bg/catalogsearch/result/?q={sku}"
    logging.info(f"→ Fetching {search_url}")
    try:
        resp = session.get(search_url, timeout=10)
        resp.raise_for_status()
    except requests.RequestException:
        logging.warning(f"← {sku}: request failed")
        return "Item doesn't exist"

    soup = BeautifulSoup(resp.text, 'html.parser')
    h1 = soup.find('h1')
    if not h1 or 'Има намерени' not in h1.get_text():
        logging.info(f"← {sku}: not found (no results header)")
        return "Item doesn't exist"

    grid = soup.find('div', class_=lambda c: c and 'grid' in c and 'place-items-center' in c)
    if grid:
        a = grid.find('a', href=True)
        if a:
            href = a['href']
            url = href if href.startswith('http') else f"https://praktis.bg{href}"
            logging.info(f"← {sku}: found {url}")
            return url

    logging.info(f"← {sku}: not found (no product link)")
    return "Item doesn't exist"


def main(input_file: str, output_dir: str):
    logging.info(f"Starting SKU URL check. Input: {input_file}")
    try:
        df = pd.read_excel(input_file, sheet_name=0, dtype=str)
    except Exception as e:
        logging.error(f"Error reading input file: {e}")
        return
    skus = df.iloc[:, 0].dropna().tolist()

    results = []
    for sku in skus:
        url = find_product_url(sku)
        results.append({'SKU': sku, 'URL': url})
        time.sleep(0.2)

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = os.path.join(output_dir, f"output_{ts}.xlsx")

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet('Results')

    ws.cell(row=1, column=1, value='SKU')
    ws.cell(row=1, column=2, value='URL')
    red_fill = get_red_fill()

    for idx, entry in enumerate(results, start=2):
        ws.cell(row=idx, column=1, value=entry['SKU'])
        cell = ws.cell(row=idx, column=2, value=entry['URL'])
        if entry['URL'].startswith('http'):
            cell.hyperlink = entry['URL']
            cell.style = 'Hyperlink'
        else:
            cell.fill = red_fill

    for col in ['A', 'B']:
        max_length = max((len(str(cell.value)) for cell in ws[col] if cell.value), default=0)
        ws.column_dimensions[col].width = max_length * 1.2 + 2

    try:
        wb.save(output_file)
        logging.info(f"Finished! Results written to: {output_file}")
    except Exception as e:
        logging.error(f"Error saving output file: {e}")


if __name__ == '__main__':
    input_file = r'C:\Users\МЕГАДОМ\Documents\url_check\primer1.xls'
    output_dir = r'C:\Users\МЕГАДОМ\Documents\url_check\output'
    main(input_file, output_dir)